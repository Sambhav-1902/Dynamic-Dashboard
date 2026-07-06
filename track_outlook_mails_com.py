# -*- coding: utf-8 -*-
"""
TechOps Mail Tracker — Office Laptop Version (Outlook COM)
==========================================================
Monitors your Outlook (Classic) inbox, logs incoming support emails to
mail_tracker.xlsx, and keeps a live dashboard updated automatically.

Differences from the personal laptop version (track_outlook_mails.py):
  - Uses Outlook COM automation (pywin32) instead of Gmail IMAP
  - No credentials required — attaches to the already-signed-in Outlook app
  - Ollama model loaded from a local GGUF file (network blocks ollama pull)
  - No username/password prompt at startup

Everything else is identical: AI models, Excel structure, status logic,
dashboard generation, retry queue.

==========================================================
SETUP — run these steps once before the first run
==========================================================

1. Install Python packages:
       pip install pywin32 openpyxl sentence-transformers torch gliner parsedatetime requests

2. Install Ollama (https://ollama.ai/download) — Windows installer.
   After installation Ollama runs as a background service automatically.

3. Load the Gemma 3 1B model into Ollama from the local GGUF file:
   a. Open Python and find where the GGUF was cached:
          from huggingface_hub import hf_hub_download
          print(hf_hub_download("bartowski/google_gemma-3-1b-it-GGUF",
                                "google_gemma-3-1b-it-Q4_K_M.gguf"))
   b. Create a plain text file called Modelfile (no extension) with:
          FROM <path printed above>
          PARAMETER temperature 0.1
   c. In a terminal, from the folder where you saved the Modelfile:
          ollama create gemma3:1b -f Modelfile
   d. Verify it worked:
          ollama run gemma3:1b "Reply YES"

4. Make sure Outlook (Classic) is open and signed in before running.

5. Update ALLOWED_EMAILS below to include your office email addresses.

==========================================================
HOW TO RUN
==========================================================

    python track_outlook_mails_com.py

The script will:
  - Load AI models (~30s on first run)
  - Connect to Outlook
  - Load any emails since the last run (or last 24hrs if fresh file)
  - Generate the dashboard
  - Poll every 15 seconds for new emails
  - Update the dashboard automatically whenever new data arrives
  - Press Ctrl+C to stop (dashboard is refreshed one final time before exit)

==========================================================
STATUS LOGIC
==========================================================

  New email from allowed sender          -> New ticket, Status = Pending
  Reply from original sender + Pending   -> Followups counter incremented
  Reply from original sender + Ongoing   -> Status stays Ongoing
  Reply from original sender + Resolved  -> Treated as a new ticket
  Reply from anyone else                 -> Gemma checks the full conversation
                                            -> Resolved or Ongoing

Category is detected once at ticket creation (never on replies).
Due dates are informational only — they do NOT change ticket status.
"""

import win32com.client
import pythoncom
import email
import requests
from email.header import decode_header
from email.utils import parseaddr, parsedate_to_datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.text import CharacterProperties, ParagraphProperties
import os
import re
import time
import parsedatetime
from datetime import datetime, timedelta
from collections import Counter, defaultdict
from gliner import GLiNER
from sentence_transformers import util as st_util

OUTPUT_FILE          = "mail_tracker.xlsx"
ALLOWED_SENDERS_FILE = "allowed_senders.txt"
POLL_INTERVAL        = 15  # seconds between inbox checks

# MAPI property tags used to read raw headers and resolve Exchange sender addresses
PR_TRANSPORT_MESSAGE_HEADERS = "http://schemas.microsoft.com/mapi/proptag/0x007D001E"
PR_SMTP_ADDRESS              = "http://schemas.microsoft.com/mapi/proptag/0x39FE001E"

OL_MAIL_ITEM    = 43  # MailItem.Class — skips meeting requests, receipts, etc.
OL_FOLDER_INBOX = 6

# Load AI models at startup
print("Loading AI models...")
deadline_model = GLiNER.from_pretrained("urchade/gliner_small-v2.1")
from sentence_transformers import SentenceTransformer
category_model = SentenceTransformer('BAAI/bge-base-en-v1.5')
print("AI models ready.\n")

# Emails from these addresses are always accepted
ALLOWED_EMAILS = [
    "sambhavsingwi@gmail.com",
    "cs1230722@iitd.ac.in",
]

ALLOWED_DOMAINS = [
    # "exlservice.com",
]

# Only process emails that were sent to or CC'd/BCC'd to this DL address.
# Emails that don't involve this address at all are ignored.
DL_ADDRESS = "support@exlservice.com"


# Allowed senders management


def load_allowed_senders():
    senders = set(e.lower() for e in ALLOWED_EMAILS)
    if os.path.exists(ALLOWED_SENDERS_FILE):
        with open(ALLOWED_SENDERS_FILE, "r", encoding="utf-8") as f:
            for line in f:
                addr = line.strip().lower()
                if addr:
                    senders.add(addr)
    return senders


def save_allowed_senders(senders):
    with open(ALLOWED_SENDERS_FILE, "w", encoding="utf-8") as f:
        for addr in sorted(senders):
            f.write(addr + "\n")


def extract_and_update_senders(msg, current_allowed):
    """Scans To/Cc/Bcc of a processed email and adds any new addresses to the allowed list."""
    new_addresses = []
    for header in ["To", "Cc", "Bcc"]:
        raw = msg.get(header, "")
        if not raw:
            continue
        for part in raw.split(","):
            _, addr = parseaddr(part.strip())
            addr = addr.lower().strip()
            if addr and "@" in addr and addr not in current_allowed:
                new_addresses.append(addr)
                current_allowed.add(addr)
    if new_addresses:
        save_allowed_senders(current_allowed)
        print(f"Added {len(new_addresses)} new sender(s) to allowed list:")
        for addr in new_addresses:
            print(f"   + {addr}")
    return current_allowed


def is_allowed(sender_email, allowed_senders):
    return sender_email.lower().strip() in allowed_senders


def is_dl_email(em):
    """Returns True if the DL address appears as a whole address in To/CC/BCC.
    Uses word-boundary matching so 'notsupport@exlservice.com' does NOT match."""
    import re as _re
    to_cc_bcc = em.get("to_cc_bcc", "")
    # Split on common separators and check each address individually
    addresses = _re.split(r'[\s;,<>]+', to_cc_bcc)
    return DL_ADDRESS.lower() in addresses


# Email parsing (COM version)


def decode_str(value):
    if not value:
        return ""
    parts = decode_header(value)
    decoded = []
    for part, charset in parts:
        if isinstance(part, bytes):
            decoded.append(part.decode(charset or "utf-8", errors="replace"))
        else:
            decoded.append(part)
    return " ".join(decoded).strip()


def strip_html(text):
    text = re.sub(r'<[^>]+>', ' ', text)
    text = re.sub(r'&nbsp;', ' ', text)
    text = re.sub(r'&amp;', '&', text)
    text = re.sub(r'&lt;', '<', text)
    text = re.sub(r'&gt;', '>', text)
    text = re.sub(r'&quot;', '"', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def get_sender_smtp_address(item):
    """Resolves the sender's SMTP address. Internal Exchange senders return an EX-format
    address by default; PR_SMTP_ADDRESS converts this to a normal email address."""
    try:
        smtp = item.Sender.PropertyAccessor.GetProperty(PR_SMTP_ADDRESS)
        if smtp:
            return smtp
    except Exception:
        pass
    try:
        if getattr(item, "SenderEmailType", "") == "SMTP":
            return item.SenderEmailAddress
    except Exception:
        pass
    return getattr(item, "SenderEmailAddress", "") or ""


def parse_email_com(item):
    """Reads an Outlook MailItem and returns the same dict structure as the Gmail version.
    Uses PR_TRANSPORT_MESSAGE_HEADERS to get real Message-ID and In-Reply-To headers."""
    try:
        raw_headers = item.PropertyAccessor.GetProperty(PR_TRANSPORT_MESSAGE_HEADERS)
    except Exception:
        raw_headers = ""

    msg = email.message_from_string(raw_headers) if raw_headers else email.message.Message()

    sender_email = get_sender_smtp_address(item)
    sender_name  = getattr(item, "SenderName", "") or sender_email
    subject      = getattr(item, "Subject", "") or "(No Subject)"

    # Parse date from header and normalize to local timezone
    date_str = None
    header_date = msg.get("Date")
    if header_date:
        try:
            parsed_dt = parsedate_to_datetime(header_date)
            if parsed_dt.tzinfo is not None:
                parsed_dt = parsed_dt.astimezone()
            date_str = parsed_dt.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            date_str = None
    if date_str is None:
        try:
            rt = item.ReceivedTime
            date_str = datetime(rt.year, rt.month, rt.day, rt.hour, rt.minute, rt.second).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            date_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    body = ""
    try:
        body = item.Body or ""
    except Exception:
        pass
    if not body.strip():
        try:
            body = strip_html(item.HTMLBody or "")
        except Exception:
            pass
    body = body.strip()[:2000]

    message_id  = msg.get("Message-ID", "").strip()
    in_reply_to = msg.get("In-Reply-To", "").strip()

    if not message_id:
        try:
            message_id = f"<entryid-{item.EntryID}@local>"
        except Exception:
            message_id = ""

    return {
        "sender_name":  sender_name or sender_email,
        "sender_email": sender_email,
        "subject":      decode_str(subject),
        "received":     date_str,
        "body":         body,
        "message_id":   message_id,
        "in_reply_to":  in_reply_to,
        "raw_msg":      msg,
        "_entry_id":    getattr(item, "EntryID", None),
        "to_cc_bcc":    " ".join(filter(None, [
            msg.get("To", ""),
            msg.get("Cc", ""),
            msg.get("Bcc", ""),
            getattr(item, "To", "") or "",
            getattr(item, "CC", "") or "",
            getattr(item, "BCC", "") or "",
        ])).lower(),
    }


def clean_subject(subject):
    cleaned = subject.strip()
    prefix_pattern = re.compile(r'^(re|fwd|fw)\s*:\s*', re.IGNORECASE)
    while prefix_pattern.match(cleaned):
        cleaned = prefix_pattern.sub('', cleaned).strip()
    return cleaned.lower()


# AI — Conversational resolution check (Gemma 3 1B via Ollama)


OLLAMA_URL   = "http://localhost:11434/api/generate"
OLLAMA_MODEL = "gemma3:1b"


def is_conversation_resolved(conversation_log, original_sender):
    """Asks Gemma 3 1B (via local Ollama) whether every issue raised by original_sender
    has been resolved across the full conversation. Returns True (Resolved) or False (Ongoing).
    Defaults to False if Ollama is unreachable."""
    prompt = f"""Below is a support ticket conversation, shown NEWEST message first. Each message shows the sender's email and their message.

Conversation:
{conversation_log}

The person who raised this ticket is: {original_sender}

Question: Has EVERY question or issue raised by {original_sender} (across all of their messages) been clearly and explicitly answered or fixed by the OTHER person's replies in this conversation?

If even ONE question or issue from {original_sender} is still unanswered or unresolved, answer NO.
Only answer YES if you can find an explicit answer/fix for EACH thing {original_sender} asked about.

Answer with ONLY one word: YES or NO."""

    try:
        response = requests.post(OLLAMA_URL, json={
            "model": OLLAMA_MODEL,
            "prompt": prompt,
            "stream": False,
            "options": {"temperature": 0.1}
        }, timeout=30)
        text = response.json()["response"].strip().upper()
        if "YES" in text and "NO" not in text:
            return True
        elif "NO" in text:
            return False
        else:
            print(f"   Unclear response from Gemma ({text!r}) — defaulting to Ongoing")
            return False
    except Exception as e:
        print(f"   Could not reach Ollama ({e}) — defaulting to Ongoing.")
        return False


# AI — Category detection


# Descriptive sentences used as category labels for semantic similarity matching
CATEGORY_LABELS = [
    'unable to login or access account, password not working, account locked',
    'payment transaction failed, duplicate charge, refund error, amount wrongly deducted, EMI payment not going through',
    'loan application status, loan approval delay, loan disbursement not credited, loan tenure or interest rate change, loan account closure',
    'wrong data in report, data mismatch, incorrect figures, reconciliation error',
    'application crash, software bug, system not working, page not loading',
    'API error, integration failure, sync timeout, third party system not responding',
    'system very slow, page timeout, high latency, unresponsive system',
    'KYC verification failed, eSign not working, document upload compliance issue',
    'how to do something, asking for information, no technical problem just a question',
]

LABEL_TO_FULL = {
    'unable to login or access account, password not working, account locked'              : 'Access and Login Issues',
    'payment transaction failed, duplicate charge, refund error, amount wrongly deducted, EMI payment not going through' : 'Payment Processing Issues',
    'loan application status, loan approval delay, loan disbursement not credited, loan tenure or interest rate change, loan account closure' : 'Loan Processing Issues',
    'wrong data in report, data mismatch, incorrect figures, reconciliation error'        : 'Data and Reporting Issues',
    'application crash, software bug, system not working, page not loading'               : 'System and Application Errors',
    'API error, integration failure, sync timeout, third party system not responding'     : 'Integration and API Issues',
    'system very slow, page timeout, high latency, unresponsive system'                   : 'Performance Issues',
    'KYC verification failed, eSign not working, document upload compliance issue'        : 'Compliance and KYC Issues',
    'how to do something, asking for information, no technical problem just a question'   : 'General Query',
}

CATEGORY_MARGIN        = 0.05  # include categories within this margin of the top score
CATEGORY_MIN_TOP_SCORE = 0.5   # if top score is below this, text is too vague — just take top 1

CATEGORY_EMBS = category_model.encode(CATEGORY_LABELS, convert_to_tensor=True)

QUESTION_WORDS = ["how to", "how do", "what is", "what are", "what's",
                  "can you tell", "can you please tell", "could you tell",
                  "please tell", "please let me know", "please explain",
                  "i want to know", "i wanted to know", "could you",
                  "would you", "is there", "are there", "do you know",
                  "when is", "when will", "when can"]

PROBLEM_WORDS = ["failed", "error", "not working", "stuck", "issue", "problem",
                 "crash", "unable", "cannot", "can't", "broken", "wrong",
                 "incorrect", "missing", "slow", "timeout", "deducted", "not loading"]


def detect_categories(body):
    """Detects one or more issue categories using BGE embeddings + margin-based matching.
    Runs once at ticket creation, never on replies. Returns a comma-separated string."""
    words = body.strip().split()
    text  = " ".join(words[:150])
    text_lower = text.lower()

    has_question = any(q in text_lower for q in QUESTION_WORDS)
    has_problem  = any(p in text_lower for p in PROBLEM_WORDS)
    if has_question and not has_problem:
        return "General Query"

    try:
        emb       = category_model.encode(text, convert_to_tensor=True)
        scores    = st_util.cos_sim(emb, CATEGORY_EMBS)[0].tolist()
        score_map = {LABEL_TO_FULL[label]: score for label, score in zip(CATEGORY_LABELS, scores)}
        top_score = max(score_map.values())

        if top_score < CATEGORY_MIN_TOP_SCORE:
            detected = [max(score_map.items(), key=lambda x: x[1])[0]]
        else:
            cutoff   = top_score - CATEGORY_MARGIN
            detected = [cat for cat, score in score_map.items() if score >= cutoff]
    except Exception as e:
        print(f"Category detection error: {e} — defaulting to General Query")
        detected = ["General Query"]

    return ", ".join(detected)


# AI — Due date extraction


_cal = parsedatetime.Calendar()

_SPECIAL_CASES = {
    "weekend"    : lambda: (datetime.now() + timedelta(days=max(1, (5 - datetime.now().weekday()) % 7))).replace(hour=17, minute=0, second=0, microsecond=0),
    "eod"        : lambda: datetime.now().replace(hour=17, minute=0, second=0, microsecond=0),
    "end of day" : lambda: datetime.now().replace(hour=17, minute=0, second=0, microsecond=0),
    "eow"        : lambda: (datetime.now() + timedelta(days=max(1, (4 - datetime.now().weekday()) % 7))).replace(hour=17, minute=0, second=0, microsecond=0),
    "end of week": lambda: (datetime.now() + timedelta(days=max(1, (4 - datetime.now().weekday()) % 7))).replace(hour=17, minute=0, second=0, microsecond=0),
}

DEADLINE_KEYWORDS = [
    "by", "before", "due", "deadline", "complete", "finish",
    "done", "fix", "resolve", "submit", "deliver", "send",
    "no later than", "needs to be", "should be", "must be",
    "have to", "need to", "expected by", "required by",
    "end of day", "end of week", "eod", "eow",
]


def find_deadline_sentence(text, entity_text):
    sentences = re.split(r'[.!?\n]', text)
    for sentence in sentences:
        if entity_text.lower() in sentence.lower():
            if any(kw in sentence.lower() for kw in DEADLINE_KEYWORDS):
                return sentence
    return None


def extract_deadline(body):
    """Extracts a due date from email body using GLiNER. Returns formatted datetime string,
    'ASAP', 'URGENT', or empty string if nothing found."""
    text       = get_first_300_words(body)
    text_lower = text.lower()

    if "asap" in text_lower:
        return "ASAP"
    if "urgent" in text_lower:
        return "URGENT"

    try:
        entities = deadline_model.predict_entities(text, ["date", "time", "duration"], threshold=0.3)
    except Exception as e:
        print(f"Due date extraction error: {e}")
        return ""

    if not entities:
        return ""

    for entity in entities:
        date_text = entity["text"].strip().lower()
        if not find_deadline_sentence(text, date_text):
            continue
        for key, fn in _SPECIAL_CASES.items():
            if key in date_text:
                return fn().strftime("%Y-%m-%d %H:%M")
        result, status = _cal.parseDT(date_text, datetime.now())
        if status != 0:
            return result.strftime("%Y-%m-%d %H:%M")

    return ""


def strip_quoted_content(body):
    """Strips quoted/forwarded content from an email, keeping only the new text."""
    text = body.strip()

    separators = [
        "________________________________",
        "-----Original Message-----",
        "-----Forwarded Message-----",
        "-------- Original Message --------",
        "-------------------------",
    ]
    for sep in separators:
        idx = text.find(sep)
        if idx > 5:
            text = text[:idx].strip()
            break

    on_wrote_pattern = re.compile(
        r'\n?On\s+(Mon|Tue|Wed|Thu|Fri|Sat|Sun)[a-z]*,.*?wrote:',
        re.IGNORECASE | re.DOTALL
    )
    match = on_wrote_pattern.search(text)
    if match and match.start() > 5:
        text = text[:match.start()].strip()

    lines = text.split("\n")
    text = "\n".join(line for line in lines if not line.strip().startswith(">")).strip()

    text = re.sub(r'CAUTION:.*?safe\.', '', text, flags=re.IGNORECASE | re.DOTALL).strip()

    return text


def get_first_300_words(body):
    """Returns the first 300 words of the new reply text (quoted content stripped)."""
    text = strip_quoted_content(body)
    return " ".join(text.split()[:300])
def extract_worth(body):
    """Extracts the first monetary amount mentioned in the email body.
    Handles formats like $100, $1,000, $1.5k, USD 500, 100 dollars, Rs 500, INR 1000.
    Returns a float (the dollar/currency amount) or 0.0 if nothing found.
    For non-USD currencies (Rs, INR) the raw number is stored as-is for comparison.
    """
    text = get_first_300_words(body)

    # Patterns: $1,000.50 / $ 1000 / USD 1,000 / 1000 USD / 1.5k dollars / Rs 500 / INR 1000
    patterns = [
        r'\$\s*([\d,]+(?:\.\d+)?)\s*[kK]?',        # $100 / $1,000 / $1.5k
        r'USD\s*([\d,]+(?:\.\d+)?)\s*[kK]?',        # USD 500
        r'([\d,]+(?:\.\d+)?)\s*USD',                 # 500 USD
        r'([\d,]+(?:\.\d+)?)\s*dollars?',            # 100 dollars
        r'Rs\.?\s*([\d,]+(?:\.\d+)?)',               # Rs 500 / Rs. 500
        r'INR\s*([\d,]+(?:\.\d+)?)',                  # INR 1000
    ]

    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            raw = match.group(1).replace(",", "")
            try:
                amount = float(raw)
                # Handle k suffix (e.g. $1.5k)
                full = match.group(0)
                if full.lower().endswith('k'):
                    amount *= 1000
                return amount
            except ValueError:
                continue

    return 0.0


# AI — Status detection helpers




# Assignee name extraction


def extract_assignee_name(sender_name, sender_email, body):
    """Extracts the assignee name. Tries sender name first, then email prefix, then GLiNER."""
    if sender_name and sender_name.strip():
        name = sender_name.strip()
        skip = ["noreply", "no-reply", "support", "team", "system",
                "notification", "alert", "admin", "info", "do-not-reply"]
        if not any(kw in name.lower() for kw in skip):
            return name

    if sender_email and "@" in sender_email:
        local_part = sender_email.split("@")[0]
        local_part = re.sub(r'^(ex_|usr_|emp_|user_|staff_)', '', local_part, flags=re.IGNORECASE)
        local_part = re.sub(r'[._-]', ' ', local_part)
        local_part = re.sub(r'\d+', '', local_part).strip()
        if local_part:
            return local_part.title()

    if body and body.strip():
        try:
            entities = deadline_model.predict_entities(get_first_300_words(body), ["person name"], threshold=0.4)
            if entities:
                return entities[0]["text"].strip().title()
        except Exception as e:
            print(f"Name extraction error: {e}")

    return ""


# Excel setup and operations


HEADERS = [
    "#", "Sender Name", "Sender Email", "Subject", "Category",
    "Received Date & Time", "Latest Reply Date & Time", "Status",
    "Ongoing Since", "Assigned To", "Followups", "Due Date", "Body", "Message-ID",
    "Worth ($)",
]

COL_NUM          = 1
COL_SENDER_NAME  = 2
COL_SENDER_EMAIL = 3
COL_SUBJECT      = 4
COL_CATEGORY     = 5
COL_RECEIVED     = 6
COL_LATEST_DATE  = 7
COL_STATUS       = 8
COL_ONGOING_SINCE = 9
COL_ASSIGNED_TO  = 10
COL_FOLLOWUPS    = 11
COL_DUE_DATE     = 12
COL_BODY         = 13
COL_MESSAGE_ID   = 14
COL_WORTH        = 15

STATUS_COLORS = {
    "Pending":  "FFF2CC",
    "Ongoing":  "DDEEFF",
    "Resolved": "E2EFDA",
}


def init_excel(filepath):
    if os.path.exists(filepath):
        print(f"Found existing file — new emails will be appended to '{filepath}'")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inbox Tracker"

    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 20
    ws.column_dimensions["J"].width = 30
    ws.column_dimensions["K"].width = 10
    ws.column_dimensions["L"].width = 20
    ws.column_dimensions["M"].width = 60
    ws.column_dimensions["N"].width = 40
    ws.column_dimensions["O"].width = 14  # Worth ($)
    ws.row_dimensions[1].height = 30

    wb.save(filepath)
    print(f"Created new tracker file: '{filepath}'")


def apply_status_color(cell, status):
    color = STATUS_COLORS.get(status, "FFFFFF")
    cell.fill      = PatternFill("solid", start_color=color)
    cell.font      = Font(name="Arial", size=10, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="top")


def get_all_message_ids(ws):
    id_to_row = {}
    for row in range(2, ws.max_row + 1):
        mid = ws.cell(row=row, column=COL_MESSAGE_ID).value
        if mid:
            id_to_row[mid.strip()] = row
    return id_to_row


def get_all_subjects(ws):
    subject_to_row = {}
    for row in range(2, ws.max_row + 1):
        subject = ws.cell(row=row, column=COL_SUBJECT).value
        if subject:
            subject_to_row[clean_subject(subject)] = row
    return subject_to_row


def find_matching_row(em, ws):
    """Matches a reply to an existing ticket — first by Message-ID, then by subject."""
    id_to_row  = get_all_message_ids(ws)
    target_row = id_to_row.get(em["in_reply_to"])
    if target_row:
        return target_row, "message-id"
    subject_to_row = get_all_subjects(ws)
    target_row     = subject_to_row.get(clean_subject(em["subject"]))
    if target_row:
        return target_row, "subject"
    return None, None


def append_new_email(em, filepath):
    """Creates a new ticket row with Status=Pending and runs AI category/due-date detection."""
    try:
        wb       = openpyxl.load_workbook(filepath)
        ws       = wb["Inbox Tracker"]
        next_row = ws.max_row + 1
        index    = next_row - 1

        alt_fill     = PatternFill("solid", start_color="DCE6F1")
        row_fill     = alt_fill if index % 2 == 0 else PatternFill()
        category     = detect_categories(em["body"])
        due_date     = extract_deadline(em["body"])
        worth        = extract_worth(em["body"])
        clean_body   = strip_quoted_content(em["body"])
        initial_body = f"[{em['received']}] {em['sender_email']}:\n{clean_body}"

        values = [
            index, em["sender_name"], em["sender_email"], em["subject"],
            category, em["received"], "",
            "Pending", "", "", 0, due_date, initial_body, em["message_id"],
            worth if worth > 0 else "",
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=next_row, column=col, value=value)
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=(col == COL_BODY))

        apply_status_color(ws.cell(row=next_row, column=COL_STATUS), "Pending")
        ws.cell(row=next_row, column=COL_FOLLOWUPS).alignment = Alignment(horizontal="center", vertical="top")

        wb.save(filepath)
        return True
    except PermissionError:
        return False


def update_existing_row(em, target_row, filepath):
    """Updates an existing ticket row based on who replied and the current status.

    - Same sender + Pending  -> increment Followups, keep Pending
    - Same sender + Ongoing  -> keep Ongoing
    - Same sender + Resolved -> return 'new_issue' (caller creates a new row)
    - Anyone else replied    -> Gemma checks full conversation -> Resolved or Ongoing
    """
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb["Inbox Tracker"]

        original_sender = (ws.cell(row=target_row, column=COL_SENDER_EMAIL).value or "").lower().strip()
        reply_sender    = em["sender_email"].lower().strip()
        current_status  = ws.cell(row=target_row, column=COL_STATUS).value or "Pending"
        is_same_sender  = (reply_sender == original_sender)

        if is_same_sender and current_status == "Resolved":
            return "new_issue"

        due_date      = extract_deadline(em["body"])
        existing_body = ws.cell(row=target_row, column=COL_BODY).value or ""
        clean_reply   = strip_quoted_content(em["body"])
        new_entry     = f"[{em['received']}] {em['sender_email']}:\n{clean_reply}"
        combined_body = f"{new_entry}\n\n---\n\n{existing_body}" if existing_body else new_entry

        if is_same_sender and current_status == "Pending":
            current_followups = ws.cell(row=target_row, column=COL_FOLLOWUPS).value or 0
            new_followups     = int(current_followups) + 1
            ws.cell(row=target_row, column=COL_FOLLOWUPS).value     = new_followups
            ws.cell(row=target_row, column=COL_FOLLOWUPS).alignment = Alignment(horizontal="center", vertical="top")
            status = "Pending"
            print(f"   Sender following up — Pending, followups now {new_followups}")

        elif is_same_sender and current_status == "Ongoing":
            status = "Ongoing"
            print(f"   Sender adding info — stays Ongoing")

        else:
            print(f"   Reply from someone else — running resolution check...")
            resolved = is_conversation_resolved(combined_body, original_sender)
            status   = "Resolved" if resolved else "Ongoing"

        combined_body = combined_body[:10000]

        ws.cell(row=target_row, column=COL_BODY).value            = combined_body
        ws.cell(row=target_row, column=COL_BODY).alignment        = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=target_row, column=COL_LATEST_DATE).value     = em["received"]
        ws.cell(row=target_row, column=COL_LATEST_DATE).alignment = Alignment(vertical="top")

        status_cell = ws.cell(row=target_row, column=COL_STATUS, value=status)
        apply_status_color(status_cell, status)

        # Record when the ticket first became Ongoing (never overwrite once set)
        if status == "Ongoing" and current_status != "Ongoing":
            if not ws.cell(row=target_row, column=COL_ONGOING_SINCE).value:
                ws.cell(row=target_row, column=COL_ONGOING_SINCE).value     = em["received"]
                ws.cell(row=target_row, column=COL_ONGOING_SINCE).alignment = Alignment(vertical="top")

        if due_date:
            ws.cell(row=target_row, column=COL_DUE_DATE).value     = due_date
            ws.cell(row=target_row, column=COL_DUE_DATE).alignment = Alignment(vertical="top")

        # Update worth if a higher value is mentioned in this reply
        new_worth = extract_worth(em["body"])
        if new_worth > 0:
            existing_worth = ws.cell(row=target_row, column=COL_WORTH).value or 0
            try:
                existing_worth = float(existing_worth)
            except (ValueError, TypeError):
                existing_worth = 0
            if new_worth > existing_worth:
                ws.cell(row=target_row, column=COL_WORTH).value     = new_worth
                ws.cell(row=target_row, column=COL_WORTH).alignment = Alignment(vertical="top")

        if not is_same_sender:
            assignee = extract_assignee_name(em["sender_name"], em["sender_email"], em["body"])
            if assignee:
                current_assigned = ws.cell(row=target_row, column=COL_ASSIGNED_TO).value or ""
                existing = [a.strip() for a in current_assigned.split(",") if a.strip()]
                if assignee not in existing:
                    existing.append(assignee)
                    ws.cell(row=target_row, column=COL_ASSIGNED_TO).value     = ", ".join(existing)
                    ws.cell(row=target_row, column=COL_ASSIGNED_TO).alignment = Alignment(vertical="top", wrap_text=True)
                    print(f"   Assigned to: {assignee}")

        wb.save(filepath)
        return True
    except PermissionError:
        return False


# Email routing


def process_email(em, retry_queue, filepath):
    """Routes an incoming email to either update an existing ticket or create a new one."""
    if em["in_reply_to"]:
        wb         = openpyxl.load_workbook(filepath)
        ws         = wb["Inbox Tracker"]
        target_row, match_type = find_matching_row(em, ws)

        if target_row:
            result = update_existing_row(em, target_row, filepath)
            if result == "new_issue":
                print(f"New issue after resolution — creating new ticket.")
                saved = append_new_email(em, filepath)
                if not saved:
                    retry_queue.append(em)
                    print(f"   Excel open — queued.")
            elif result is True:
                print(f"Reply received — ticket updated (matched by {match_type})")
                print(f"   Subject    : {em['subject']}")
                print(f"   Replied at : {em['received']}\n")
            else:
                retry_queue.append(em)
                print(f"Excel open — reply queued: '{em['subject']}'\n")
        else:
            print(f"Reply with no matching ticket — saving as new.")
            saved = append_new_email(em, filepath)
            if not saved:
                retry_queue.append(em)
    else:
        saved = append_new_email(em, filepath)
        if not saved:
            retry_queue.append(em)
            print(f"Excel open — new ticket queued.")
        print(f"{'New ticket saved' if saved else 'Queued'}: '{em['subject']}' from {em['sender_email']}\n")


# Outlook connection helpers


def connect_outlook():
    """Connects to the already-running Outlook session via COM. No credentials needed."""
    outlook   = win32com.client.Dispatch("Outlook.Application")
    namespace = outlook.GetNamespace("MAPI")
    inbox     = namespace.GetDefaultFolder(OL_FOLDER_INBOX)
    return namespace, inbox


def get_inbox_entry_ids(inbox):
    """Returns the set of EntryIDs for all mail items in the inbox."""
    try:
        ids   = set()
        items = inbox.Items
        for item in items:
            try:
                if item.Class == OL_MAIL_ITEM:
                    ids.add(item.EntryID)
            except Exception:
                continue
        return ids
    except Exception as e:
        print(f"Could not reach Outlook inbox: {e}")
        return None


def fetch_item_by_entry_id(namespace, entry_id):
    """Fetches and parses a single mail item by EntryID."""
    try:
        item = namespace.GetItemFromID(entry_id)
        if item.Class == OL_MAIL_ITEM:
            return parse_email_com(item)
    except Exception as e:
        print(f"Could not fetch email: {e}")
    return None


def fetch_items_since(inbox, since_dt):
    """Fetches all mail items received after since_dt, sorted oldest-first with
    original tickets before replies (prevents reply-before-ticket race conditions)."""
    results = []
    try:
        items = inbox.Items
        items.Sort("[ReceivedTime]", False)
        restrict_str = since_dt.strftime("[ReceivedTime] > '%m/%d/%Y %H:%M %p'")
        try:
            filtered = items.Restrict(restrict_str)
        except Exception:
            filtered = items

        for item in filtered:
            try:
                if item.Class != OL_MAIL_ITEM:
                    continue
                em = parse_email_com(item)
                try:
                    em_dt = datetime.strptime(em["received"], "%Y-%m-%d %H:%M:%S")
                except Exception:
                    em_dt = since_dt
                if em_dt > since_dt:
                    results.append((em["_entry_id"], em, em_dt))
            except Exception:
                continue
    except Exception as e:
        print(f"Could not fetch historical emails: {e}")
        return []

    results.sort(key=lambda x: x[2])
    no_reply  = [(eid, em, dt) for eid, em, dt in results if not em["in_reply_to"]]
    has_reply = [(eid, em, dt) for eid, em, dt in results if em["in_reply_to"]]
    no_reply.sort(key=lambda x: x[2])
    has_reply.sort(key=lambda x: x[2])
    return [(eid, em) for eid, em, _ in no_reply + has_reply]


# Historical load


def get_last_record_time(filepath):
    """Returns the latest timestamp in the Excel file, or None if empty/missing."""
    if not os.path.exists(filepath):
        return None
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb["Inbox Tracker"]
        if ws.max_row <= 1:
            return None

        latest_dt = None
        for row in range(2, ws.max_row + 1):
            for col in [COL_RECEIVED, COL_LATEST_DATE]:
                val = ws.cell(row=row, column=col).value
                if not val:
                    continue
                try:
                    dt = datetime.strptime(str(val), "%Y-%m-%d %H:%M:%S")
                    if latest_dt is None or dt > latest_dt:
                        latest_dt = dt
                except Exception:
                    continue
        return latest_dt
    except Exception as e:
        print(f"Could not read last record time: {e}")
        return None


def load_historical_emails_com(inbox, allowed_senders, retry_queue):
    """Loads emails since the last recorded timestamp (or last 24hrs if the file is empty)."""
    last_time = get_last_record_time(OUTPUT_FILE)

    if last_time is None:
        since_dt = datetime.now() - timedelta(hours=24)
        print(f"No existing data — loading emails from last 24 hours ({since_dt.strftime('%Y-%m-%d %H:%M:%S')})...")
    else:
        since_dt = last_time
        print(f"Existing data found — loading since {since_dt.strftime('%Y-%m-%d %H:%M:%S')}...")

    emails = fetch_items_since(inbox, since_dt)

    if not emails:
        print("No new historical emails to load.")
        return allowed_senders

    print(f"Found {len(emails)} historical email(s). Processing oldest first...")
    print()

    for i, (eid, em) in enumerate(emails, 1):
        print(f"[{i}/{len(emails)}] {em['subject']} — from {em['sender_email']}")
        if not is_allowed(em["sender_email"], allowed_senders) and not is_dl_email(em):
            print(f"   Ignored (sender unknown and DL not in To/CC/BCC)")
            continue
        allowed_senders = extract_and_update_senders(em["raw_msg"], allowed_senders)
        process_email(em, retry_queue, OUTPUT_FILE)

    print(f"Historical load complete — {len(emails)} email(s) processed.\n")
    return allowed_senders


def check_due_date_and_resolve(filepath):
    pass


# Dashboard generation


DASH_SHEET          = "Dashboard"
DASH_STATUSES       = ["Pending", "Ongoing", "Resolved"]
DASH_STATUS_COLORS  = {"Pending": "FFC000", "Ongoing": "5B9BD5", "Resolved": "70AD47"}

DASH_TITLE_FONT        = Font(bold=True, size=16, name="Arial", color="1F4E79")
DASH_SUBTITLE_FONT     = Font(italic=True, size=9, name="Arial", color="888888")
DASH_SECTION_FONT      = Font(bold=True, size=12, name="Arial", color="FFFFFF")
DASH_SECTION_FILL      = PatternFill("solid", start_color="2E75B6")
DASH_LABEL_FONT        = Font(bold=True, size=10, name="Arial")
DASH_VALUE_FONT        = Font(size=10, name="Arial")
DASH_TABLE_HEADER_FONT = Font(bold=True, size=10, name="Arial", color="FFFFFF")
DASH_TABLE_HEADER_FILL = PatternFill("solid", start_color="44546A")
DASH_URGENT_FILL       = PatternFill("solid", start_color="FCE4D6")


def _dash_set_title(title_holder, size_pt=14, bold=True, color="1F1F1F"):
    """Sets explicit font size and overlay=False on a chart/axis title.
    Required for openpyxl 3.1.4+ where titles are tiny and overlapping in real Excel."""
    if title_holder.title is None or title_holder.title.tx is None:
        return
    para = title_holder.title.tx.rich.p[0]
    cp = CharacterProperties(sz=size_pt * 100, b=bold, solidFill=color)
    para.pPr = ParagraphProperties(defRPr=cp)
    for run in para.r:
        run.rPr = cp
    title_holder.title.overlay = False


def _dash_section_header(ws, row, text, span=6):
    cell = ws.cell(row=row, column=2, value=text)
    cell.font = DASH_SECTION_FONT
    cell.fill = DASH_SECTION_FILL
    for c in range(2, 2 + span):
        ws.cell(row=row, column=c).fill = DASH_SECTION_FILL
    return row + 1


def _dash_parse_dt(value, fmt="%Y-%m-%d %H:%M:%S"):
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    try:
        return datetime.strptime(str(value), fmt)
    except Exception:
        return None


def _dash_read_tickets(data_ws):
    tickets = []
    for row in range(2, data_ws.max_row + 1):
        num = data_ws.cell(row=row, column=COL_NUM).value
        if num is None:
            continue
        raw_worth = data_ws.cell(row=row, column=COL_WORTH).value
        try:
            worth = float(raw_worth) if raw_worth else 0.0
        except (ValueError, TypeError):
            worth = 0.0
        tickets.append({
            "num":           num,
            "sender_name":   data_ws.cell(row=row, column=COL_SENDER_NAME).value or "",
            "sender_email":  data_ws.cell(row=row, column=COL_SENDER_EMAIL).value or "",
            "subject":       data_ws.cell(row=row, column=COL_SUBJECT).value or "",
            "category":      data_ws.cell(row=row, column=COL_CATEGORY).value or "",
            "received":      data_ws.cell(row=row, column=COL_RECEIVED).value or "",
            "latest":        data_ws.cell(row=row, column=COL_LATEST_DATE).value or "",
            "status":        data_ws.cell(row=row, column=COL_STATUS).value or "",
            "ongoing_since": data_ws.cell(row=row, column=COL_ONGOING_SINCE).value or "",
            "assigned_to":   data_ws.cell(row=row, column=COL_ASSIGNED_TO).value or "",
            "followups":     data_ws.cell(row=row, column=COL_FOLLOWUPS).value or 0,
            "due_date":      data_ws.cell(row=row, column=COL_DUE_DATE).value or "",
            "worth":         worth,
        })
    return tickets


def _dash_status_counts(tickets):
    counts = Counter(t["status"] for t in tickets)
    return {s: counts.get(s, 0) for s in DASH_STATUSES}


def _dash_category_breakdown(tickets, categories):
    """Counts Pending/Ongoing tickets per category and sums their worth.
    Multi-label tickets count toward each category."""
    breakdown = {cat: {"Pending": 0, "Ongoing": 0, "worth": 0.0} for cat in categories}
    for t in tickets:
        if t["status"] not in ("Pending", "Ongoing"):
            continue
        cats = [c.strip() for c in (t["category"] or "").split(",") if c.strip()]
        for cat in cats:
            if cat in breakdown:
                breakdown[cat][t["status"]] += 1
                breakdown[cat]["worth"] += t.get("worth", 0.0)
    return breakdown


WORTH_BUCKETS = [
    (0,    100,   "$0–100"),
    (100,  1000,  "$100–1K"),
    (1000, 5000,  "$1K–5K"),
    (5000, None,  "$5K+"),
]

# Colors for the 4 worth brackets (green, yellow, light blue, red)
WORTH_BRACKET_COLORS = ["70AD47", "FFC000", "9DC3E6", "FF0000"]


def _dash_worth_buckets(tickets):
    """Counts all tickets by worth bracket."""
    counts = {label: 0 for _, _, label in WORTH_BUCKETS}
    for t in tickets:
        w = t.get("worth", 0.0)
        for lo, hi, label in WORTH_BUCKETS:
            if hi is None:
                if w >= lo:
                    counts[label] += 1
                    break
            elif lo <= w < hi:
                counts[label] += 1
                break
    return counts


def _dash_cat_worth_bracket_breakdown(tickets, categories):
    """For each category, counts how many tickets (all statuses) fall in each
    worth bracket. Returns {category: {bracket_label: count}}."""
    breakdown = {cat: {label: 0 for _, _, label in WORTH_BUCKETS}
                 for cat in categories}
    for t in tickets:
        w = t.get("worth", 0.0)
        cats = [c.strip() for c in (t["category"] or "").split(",") if c.strip()]
        for cat in cats:
            if cat not in breakdown:
                continue
            for lo, hi, label in WORTH_BUCKETS:
                if hi is None:
                    if w >= lo:
                        breakdown[cat][label] += 1
                        break
                elif lo <= w < hi:
                    breakdown[cat][label] += 1
                    break
    return breakdown


def _dash_assignee_breakdown(tickets):
    """Counts Pending/Ongoing tickets per assignee. Multi-assignee tickets count toward each.
    Returns top 12 by total, with unassigned tickets grouped under 'Unassigned'."""
    breakdown = defaultdict(lambda: {"Pending": 0, "Ongoing": 0})
    for t in tickets:
        if t["status"] not in ("Pending", "Ongoing"):
            continue
        names = [a.strip() for a in (t["assigned_to"] or "").split(",") if a.strip()] or ["Unassigned"]
        for name in names:
            breakdown[name][t["status"]] += 1
    sorted_items = sorted(breakdown.items(), key=lambda kv: -(kv[1]["Pending"] + kv[1]["Ongoing"]))
    return dict(sorted_items[:12])


def _dash_ongoing_list(tickets, now):
    """Returns Ongoing tickets sorted by days-in-Ongoing descending."""
    ongoing = [t for t in tickets if t["status"] == "Ongoing"]
    enriched = []
    for t in ongoing:
        since_dt = _dash_parse_dt(t["ongoing_since"])
        days = (now - since_dt).days if since_dt else 0
        enriched.append({**t, "days_ongoing": days})
    enriched.sort(key=lambda t: -t["days_ongoing"])
    return enriched


def update_dashboard(filepath):
    """Rebuilds the Dashboard sheet as a static snapshot. Returns True on success."""
    try:
        wb = openpyxl.load_workbook(filepath)
    except PermissionError:
        print("Could not update dashboard — Excel is open.")
        return False
    except Exception as e:
        print(f"Could not update dashboard: {e}")
        return False

    if "Inbox Tracker" not in wb.sheetnames:
        print("Could not update dashboard — 'Inbox Tracker' sheet not found.")
        return False

    data_ws    = wb["Inbox Tracker"]
    tickets    = _dash_read_tickets(data_ws)
    now        = datetime.now()
    categories = list(LABEL_TO_FULL.values())

    # All chart source data lives on a hidden sheet so nothing leaks
    # onto the visible Dashboard as stray tables.
    DATA_SHEET = "_ChartData"
    if DATA_SHEET in wb.sheetnames:
        del wb[DATA_SHEET]
    dws = wb.create_sheet(DATA_SHEET)
    dws.sheet_state = "hidden"

    if DASH_SHEET in wb.sheetnames:
        del wb[DASH_SHEET]
    ws = wb.create_sheet(DASH_SHEET, 0)
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation   = "landscape"
    ws.page_setup.fitToWidth    = 1
    ws.page_setup.fitToHeight   = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    status_counts      = _dash_status_counts(tickets)
    cat_breakdown       = _dash_category_breakdown(tickets, categories)
    assignee_breakdown  = _dash_assignee_breakdown(tickets)
    ongoing_list        = _dash_ongoing_list(tickets, now)
    worth_buckets       = _dash_worth_buckets(tickets)
    cat_worth_brackets  = _dash_cat_worth_bracket_breakdown(tickets, categories)

    total           = len(tickets)
    resolution_rate = (status_counts["Resolved"] / total) if total else 0
    avg_followups   = (sum(t["followups"] for t in tickets) / total) if total else 0

    bucket_labels = [label for _, _, label in WORTH_BUCKETS]

    # ------------------------------------------------------------------
    # Write all chart data to _ChartData (hidden sheet)
    # ------------------------------------------------------------------

    # Pie: col A (status labels), col B (counts)
    PIE_ROW = 1
    for i, s in enumerate(DASH_STATUSES):
        dws.cell(row=PIE_ROW + i, column=1, value=s)
        dws.cell(row=PIE_ROW + i, column=2, value=status_counts[s])
    pie_end = PIE_ROW + len(DASH_STATUSES) - 1

    # Category × worth-bracket: col D (category), cols E-H (one per bracket)
    # Row 1 = header, rows 2+ = one row per category
    CAT_BRKT_ROW = 1
    dws.cell(row=CAT_BRKT_ROW, column=4, value="Category")
    for bi, lbl in enumerate(bucket_labels):
        dws.cell(row=CAT_BRKT_ROW, column=5 + bi, value=lbl)
    for ci, cat in enumerate(categories):
        r = CAT_BRKT_ROW + 1 + ci
        dws.cell(row=r, column=4, value=cat)
        for bi, lbl in enumerate(bucket_labels):
            dws.cell(row=r, column=5 + bi, value=cat_worth_brackets[cat][lbl])
    cat_brkt_end = CAT_BRKT_ROW + len(categories)
    # col 5 = first bracket, col 5+len-1 = last bracket
    cat_brkt_last_col = 5 + len(bucket_labels) - 1

    # Assignee: col J (name), K (Pending), L (Ongoing)
    ASGN_ROW = 1
    dws.cell(row=ASGN_ROW, column=10, value="Assignee")
    dws.cell(row=ASGN_ROW, column=11, value="Pending")
    dws.cell(row=ASGN_ROW, column=12, value="Ongoing")
    for i, (name, counts) in enumerate(assignee_breakdown.items()):
        r = ASGN_ROW + 1 + i
        dws.cell(row=r, column=10, value=name)
        dws.cell(row=r, column=11, value=counts["Pending"])
        dws.cell(row=r, column=12, value=counts["Ongoing"])
    asgn_end = ASGN_ROW + len(assignee_breakdown)

    # Worth-bucket totals: col N (label), O (count)
    WBKT_ROW = 1
    dws.cell(row=WBKT_ROW, column=14, value="Worth Range")
    dws.cell(row=WBKT_ROW, column=15, value="Tickets")
    for i, (_, _, label) in enumerate(WORTH_BUCKETS):
        r = WBKT_ROW + 1 + i
        dws.cell(row=r, column=14, value=label)
        dws.cell(row=r, column=15, value=worth_buckets[label])
    wbkt_end = WBKT_ROW + len(WORTH_BUCKETS)

    # ------------------------------------------------------------------
    # Dashboard layout
    # ------------------------------------------------------------------

    ws["B2"] = "TechOps Ticket Dashboard"
    ws["B2"].font = DASH_TITLE_FONT
    ws["B3"] = f"Generated: {now.strftime('%Y-%m-%d %H:%M')}"
    ws["B3"].font = DASH_SUBTITLE_FONT

    # Section 1: Overview
    row = 5
    row = _dash_section_header(ws, row, "Overview")
    for label, value, numfmt in [
        ("Total Tickets",          total,                     None),
        ("Pending",                status_counts["Pending"],  None),
        ("Ongoing",                status_counts["Ongoing"],  None),
        ("Resolved",               status_counts["Resolved"], None),
        ("Resolution Rate",        resolution_rate,           "0.0%"),
        ("Avg Followups / Ticket", avg_followups,             "0.00"),
    ]:
        ws.cell(row=row, column=2, value=label).font = DASH_LABEL_FONT
        c = ws.cell(row=row, column=4, value=value)
        c.font = DASH_VALUE_FONT
        if numfmt:
            c.number_format = numfmt
        row += 1
    row += 1

    # ------------------------------------------------------------------
    # Section 2: Pie chart (col B) + Category-worth-bracket bar (col F)
    # One empty column gap between them (pie ends ~col E, bar starts col F).
    # ------------------------------------------------------------------
    row = _dash_section_header(ws, row, "Status & Category Breakdown", span=16)
    charts_row = row

    # Pie chart
    pie = PieChart()
    pie.title = "Tickets by Status"
    pie.add_data(
        Reference(dws, min_col=2, min_row=PIE_ROW, max_row=pie_end),
        titles_from_data=False,
    )
    pie.set_categories(
        Reference(dws, min_col=1, min_row=PIE_ROW, max_row=pie_end),
    )
    pie.height = 9
    pie.width  = 12
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showVal       = False
    pie.dataLabels.showPercent   = True
    pie.dataLabels.showCatName   = True
    pie.dataLabels.showLegendKey = False
    pie.dataLabels.showSerName   = False
    _dash_set_title(pie, size_pt=14)
    ws.add_chart(pie, f"B{charts_row}")

    # Category × worth-bracket clustered bar chart
    # 4 series (one per bracket), 9 groups (one per category)
    # Anchored at col F — one column gap after the pie chart
    cat_brkt_bar = BarChart()
    cat_brkt_bar.type     = "col"
    cat_brkt_bar.grouping = "clustered"
    cat_brkt_bar.title    = "Tickets by Category & Worth ($)"
    cat_brkt_bar.y_axis.title = "Number of Tickets"
    cat_brkt_bar.y_axis.numFmt    = "0"
    cat_brkt_bar.y_axis.majorUnit = 1
    cat_brkt_bar.add_data(
        Reference(dws, min_col=5, max_col=cat_brkt_last_col,
                  min_row=CAT_BRKT_ROW, max_row=cat_brkt_end),
        titles_from_data=True,
    )
    cat_brkt_bar.set_categories(
        Reference(dws, min_col=4,
                  min_row=CAT_BRKT_ROW + 1, max_row=cat_brkt_end),
    )
    for bi, color in enumerate(WORTH_BRACKET_COLORS):
        cat_brkt_bar.series[bi].graphicalProperties.solidFill = color
    cat_brkt_bar.height = 9
    cat_brkt_bar.width  = 22
    cat_brkt_bar.x_axis.textRotation = -30
    cat_brkt_bar.x_axis.delete = False
    cat_brkt_bar.y_axis.delete = False
    from openpyxl.chart.legend import Legend
    cat_brkt_bar.legend = Legend()
    cat_brkt_bar.legend.position = "r"
    _dash_set_title(cat_brkt_bar,        size_pt=14)
    _dash_set_title(cat_brkt_bar.y_axis, size_pt=10)
    ws.add_chart(cat_brkt_bar, f"F{charts_row}")

    row = charts_row + 19

    # ------------------------------------------------------------------
    # Section 3: Assignee bar chart
    # ------------------------------------------------------------------
    row = _dash_section_header(ws, row, "Tickets by Assignee (Pending vs Ongoing)")
    assignee_chart_row = row

    assignee_bar = BarChart()
    assignee_bar.type     = "col"
    assignee_bar.grouping = "clustered"
    assignee_bar.title    = "Tickets by Assignee"
    assignee_bar.y_axis.title = "Number of Tickets"
    assignee_bar.y_axis.numFmt    = "0"
    assignee_bar.y_axis.majorUnit = 1
    assignee_bar.add_data(
        Reference(dws, min_col=11, max_col=12,
                  min_row=ASGN_ROW, max_row=asgn_end),
        titles_from_data=True,
    )
    assignee_bar.set_categories(
        Reference(dws, min_col=10,
                  min_row=ASGN_ROW + 1, max_row=asgn_end),
    )
    assignee_bar.series[0].graphicalProperties.solidFill = DASH_STATUS_COLORS["Pending"]
    assignee_bar.series[1].graphicalProperties.solidFill = DASH_STATUS_COLORS["Ongoing"]
    assignee_bar.height = 9
    assignee_bar.width  = 18
    assignee_bar.x_axis.textRotation = -30
    assignee_bar.x_axis.delete = False
    assignee_bar.y_axis.delete = False
    _dash_set_title(assignee_bar,        size_pt=14)
    _dash_set_title(assignee_bar.y_axis, size_pt=10)
    ws.add_chart(assignee_bar, f"B{assignee_chart_row}")
    row = assignee_chart_row + 20

    # ------------------------------------------------------------------
    # Section 4: Standalone ticket worth distribution bar chart
    # ------------------------------------------------------------------
    row = _dash_section_header(ws, row, "Ticket Worth Distribution ($)")
    worth_chart_row = row

    worth_bar = BarChart()
    worth_bar.type     = "col"
    worth_bar.grouping = "clustered"
    worth_bar.title    = "Tickets by Worth ($)"
    worth_bar.y_axis.title = "Number of Tickets"
    worth_bar.y_axis.numFmt    = "0"
    worth_bar.y_axis.majorUnit = 1
    worth_bar.add_data(
        Reference(dws, min_col=15,
                  min_row=WBKT_ROW, max_row=wbkt_end),
        titles_from_data=True,
    )
    worth_bar.set_categories(
        Reference(dws, min_col=14,
                  min_row=WBKT_ROW + 1, max_row=wbkt_end),
    )
    worth_bar.series[0].graphicalProperties.solidFill = "2E75B6"
    worth_bar.height = 9
    worth_bar.width  = 18
    worth_bar.x_axis.delete = False
    worth_bar.y_axis.delete = False
    worth_bar.legend = None
    _dash_set_title(worth_bar,        size_pt=14)
    _dash_set_title(worth_bar.y_axis, size_pt=10)
    ws.add_chart(worth_bar, f"B{worth_chart_row}")
    row = worth_chart_row + 20

    # ------------------------------------------------------------------
    # Section 5: Ongoing tickets list (with Worth column)
    # ------------------------------------------------------------------
    row = _dash_section_header(ws, row,
        f"Ongoing Tickets ({len(ongoing_list)}) — Sorted by Days Ongoing", span=8)
    for i, h in enumerate(["Ticket #", "Subject", "Assignee", "Category",
                            "Days Ongoing", "Worth ($)", "Sender"]):
        c = ws.cell(row=row, column=2 + i, value=h)
        c.font = DASH_TABLE_HEADER_FONT
        c.fill = DASH_TABLE_HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    thin   = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if not ongoing_list:
        ws.cell(row=row, column=2,
                value="No tickets are currently Ongoing.").font = DASH_VALUE_FONT
    else:
        for t in ongoing_list:
            worth_disp   = f"${t['worth']:.2f}" if t.get("worth", 0) > 0 else "—"
            values       = [t["num"], t["subject"], t["assigned_to"] or "Unassigned",
                            t["category"], t["days_ongoing"], worth_disp, t["sender_email"]]
            long_ongoing = t["days_ongoing"] >= 14
            row_fill     = DASH_URGENT_FILL if long_ongoing else PatternFill()
            for col_offset, val in enumerate(values):
                c = ws.cell(row=row, column=2 + col_offset, value=val)
                c.font   = Font(name="Arial", size=10, bold=long_ongoing)
                c.fill   = row_fill
                c.border = border
                if col_offset in (0, 4, 5):
                    c.alignment = Alignment(horizontal="center", vertical="top")
                else:
                    c.alignment = Alignment(vertical="top", wrap_text=(col_offset == 1))
            row += 1

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 13
    ws.column_dimensions["H"].width = 13

    # Restore Inbox Tracker as active sheet
    wb.active = wb.sheetnames.index("Inbox Tracker")

    try:
        wb.save(filepath)
    except PermissionError:
        print("Could not save dashboard — Excel is open.")
        return False

    # Patch the category-worth chart XML directly to inject a manualLayout
    # on the plot area — openpyxl sets the object correctly in memory but
    # doesn't serialise it to XML, so the legend still overlaps the bars.
    # The patch constrains the plot area to 75% of chart width, leaving
    # the right 25% clear for the legend.
    try:
        import zipfile, shutil, os, re as _re
        tmp = filepath + ".tmp"
        with zipfile.ZipFile(filepath, 'r') as zin, \
             zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename.endswith('.xml') and b'Category' in data and b'Worth' in data \
                        and data.count(b'<ser>') >= 4:
                    # This is the category-worth chart — inject manualLayout
                    xml = data.decode('utf-8')
                    layout_xml = (
                        '<layout>'
                        '<manualLayout>'
                        '<layoutTarget val="inner"/>'
                        '<xMode val="factor"/>'
                        '<yMode val="factor"/>'
                        '<x val="0"/>'
                        '<y val="0"/>'
                        '<w val="0.75"/>'
                        '<h val="1"/>'
                        '</manualLayout>'
                        '</layout>'
                    )
                    # Insert after <plotArea> opening tag
                    xml = xml.replace('<plotArea>', '<plotArea>' + layout_xml, 1)
                    data = xml.encode('utf-8')
                zout.writestr(item, data)
        shutil.move(tmp, filepath)
    except Exception as e:
        print(f"Note: could not patch chart layout ({e}) — legend may overlap in some renderers.")

    return True

# Main listener loop


def listen():
    print("\nConnecting to Outlook...")
    pythoncom.CoInitialize()

    try:
        namespace, inbox = connect_outlook()
    except Exception as e:
        print(f"Could not connect to Outlook: {e}")
        print("Make sure Outlook (Classic) is open and signed in.")
        return

    known_ids       = get_inbox_entry_ids(inbox)
    poll_count      = 0
    retry_queue     = []
    allowed_senders = load_allowed_senders()

    if known_ids is None:
        print("Could not read inbox. Please check Outlook is running.")
        return

    print(f"Connected. {len(known_ids)} existing email(s) in inbox will be ignored.")
    print(f"Loaded {len(allowed_senders)} allowed sender(s).")

    print("Checking for historical emails...")
    allowed_senders = load_historical_emails_com(inbox, allowed_senders, retry_queue)

    print("Updating dashboard...")
    if update_dashboard(OUTPUT_FILE):
        print("Dashboard updated.\n")
    else:
        print("Could not update dashboard right now.\n")

    print(f"Listening — checking every {POLL_INTERVAL}s. Press Ctrl+C to stop.\n")

    while True:
        try:
            time.sleep(POLL_INTERVAL)
            poll_count += 1

            # Retry any emails that failed because Excel was open
            if retry_queue:
                still_failed = []
                for em in retry_queue:
                    if em["in_reply_to"]:
                        wb            = openpyxl.load_workbook(OUTPUT_FILE)
                        ws            = wb["Inbox Tracker"]
                        target_row, _ = find_matching_row(em, ws)
                        if target_row:
                            result = update_existing_row(em, target_row, OUTPUT_FILE)
                            if result == "new_issue":
                                saved = append_new_email(em, OUTPUT_FILE)
                                if not saved:
                                    still_failed.append(em)
                                else:
                                    print(f"Saved new issue from queue: '{em['subject']}'")
                            elif result is True:
                                print(f"Updated from queue: '{em['subject']}'")
                            else:
                                still_failed.append(em)
                        else:
                            saved = append_new_email(em, OUTPUT_FILE)
                            if saved:
                                print(f"Saved from queue: '{em['subject']}'")
                            else:
                                still_failed.append(em)
                    else:
                        saved = append_new_email(em, OUTPUT_FILE)
                        if saved:
                            print(f"Saved from queue: '{em['subject']}'")
                        else:
                            still_failed.append(em)
                retry_queue = still_failed
                if retry_queue:
                    print(f"Excel still open — {len(retry_queue)} email(s) waiting.")
                else:
                    update_dashboard(OUTPUT_FILE)

            check_due_date_and_resolve(OUTPUT_FILE)

            current_ids = get_inbox_entry_ids(inbox)
            if current_ids is None:
                print(f"Could not reach Outlook this cycle — retrying in {POLL_INTERVAL}s...")
                continue

            new_ids = current_ids - known_ids

            if new_ids:
                for eid in new_ids:
                    em = fetch_item_by_entry_id(namespace, eid)
                    if em:
                        if not is_allowed(em["sender_email"], allowed_senders) and not is_dl_email(em):
                            print(f"Ignored: {em['sender_email']} (sender unknown and DL not in To/CC/BCC)")
                            continue
                        allowed_senders = extract_and_update_senders(em["raw_msg"], allowed_senders)
                        process_email(em, retry_queue, OUTPUT_FILE)
                known_ids = current_ids
                update_dashboard(OUTPUT_FILE)
            else:
                if poll_count % 4 == 0:
                    queued_note = f"  ({len(retry_queue)} queued)" if retry_queue else ""
                    print(f"Listening...  {datetime.now().strftime('%H:%M:%S')}{queued_note}")

        except KeyboardInterrupt:
            if retry_queue:
                print(f"\nNote: {len(retry_queue)} email(s) unsaved (Excel was open):")
                for em in retry_queue:
                    print(f"   - '{em['subject']}' from {em['sender_email']}")
            print("\nUpdating dashboard before stopping...")
            if update_dashboard(OUTPUT_FILE):
                print("Dashboard updated.")
            else:
                print("Could not update dashboard.")
            print("Stopped.")
            break


def main():
    print("TechOps Mail Tracker — Office Laptop (Outlook COM)\n")
    init_excel(OUTPUT_FILE)
    listen()


if __name__ == "__main__":
    main()
